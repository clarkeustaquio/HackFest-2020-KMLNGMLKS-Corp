import pytz
import json
import datetime

from rest_framework import serializers, status
from rest_framework.response import Response
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.authtoken.models import Token

from django.conf import settings

from .serializers import AuthorizeUserSerializer, AnnouncementSerializer
from .models import AuthorizeUser, Announcements

# Twilio
from twilio.rest import Client, TwilioException
from twilio.twiml.messaging_response import MessagingResponse

from api import db

from openpyxl import load_workbook

client = Client(settings.TWILIO_SID, settings.TWILIO_TOKEN)
    
@api_view(['POST'])
def create_account(request):
    serializer = AuthorizeUserSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def authorize_logout(request):
    Token.objects.get(user=request.user).delete()
    return Response({ 'status': 'OK' }, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def authorize_token(request):
    username = request.user.username

    try:
        user = AuthorizeUser.objects.get(username=username)
    except AuthorizeUser.DoesNotExist:
        return Response({
            'status': 'Authorize User Does Not Exist.'
        }, status=status.HTTP_400_BAD_REQUEST)

    is_admin = user.is_staff

    return Response({
        'is_admin': is_admin
    }, status=status.HTTP_200_OK)

def get_subscribers(location):
    list_numbers = list()
    cloud_datas = db.collection('Subscribers').get()

    count = 1
    for data in cloud_datas:
        if location == data.get('location'):
            try:
                list_numbers.append({ 
                    'count': count,
                    'phone_number': data.get('phoneNumber'),
                    'location': data.get('location')
                })

                count += 1
            except KeyError:
                pass
    
    return list_numbers

def admin_subscribers():
    list_numbers = list()
    cloud_datas = db.collection('Subscribers').get()

    count = 1
    for data in cloud_datas:
        try:
            list_numbers.append({ 
                'count': count,
                'phone_number': data.get('phoneNumber'),
                'location': data.get('location')
            })

            count += 1
        except KeyError:
            pass
    
    return list_numbers

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_people(request):
    username = request.user.username

    try:
        user = AuthorizeUser.objects.get(username=username)
    except AuthorizeUser.DoesNotExist:
        return Response({
            'status': 'Authorize User Does Not Exist.'
        }, status=status.HTTP_400_BAD_REQUEST)

    place = user.place
    list_numbers = get_subscribers(place)
    
    return Response({
        'list_numbers': list_numbers,
        'place': place
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def send_announcement(request):
    username = request.user.username
    
    try:
        user = AuthorizeUser.objects.get(username=username)
    except AuthorizeUser.DoesNotExist:
        return Response({
            'status': 'Authorize User Does Not Exist.'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
         announcement = request.data['announcement']
    except KeyError:
        return Response({
            'announcement': 'This field is required.'
        })

    place = user.place
    list_numbers = get_subscribers(place)

    current_date = datetime.datetime.now().replace(
        tzinfo=pytz.UTC).astimezone(pytz.timezone('Asia/Manila')).strftime('%B %d, %Y %I:%M:%S %p')

    full_name = user.last_name.title() + ', ' + user.first_name.title()
    announcement_message = '{} - Daily Bayanihan News. \n\nFrom: {} \n\n{}'.format(current_date, full_name, announcement)

    Announcements.objects.create(
        user=user,
        announcement=announcement_message
    )

    for number in list_numbers:
        try:
            client.messages.create(
                body=announcement_message[:1600], 
                from_=settings.TWILIO_NUMBER, to=number['phone_number']
            )
        except TwilioException:
            pass
   
    return Response({
        'status': 'Announcement was successfully sent!'
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def add_subscriber(request):
    username = request.user.username
    
    try:
        user = AuthorizeUser.objects.get(username=username)
    except AuthorizeUser.DoesNotExist:
        return Response({
            'status': 'Authorize User Does Not Exist.'
        }, status=status.HTTP_400_BAD_REQUEST)

    cloud_db = db.collection('Subscribers')

    try:
        phone_number = request.data['phone_number']
    except KeyError:
        return Response({
            'phone_number': 'This field is required.'
        })

    try:
        location = request.data['location']
    except KeyError:
        return Response({
            'location': 'This field is required.'
        })

    # location = user.place

    for cloud in cloud_db.get():
        try:
            if cloud.get('phoneNumber') == phone_number:
                return Response({
                    'status': 'This number already exist!'
                }, status=status.HTTP_400_BAD_REQUEST)
        except KeyError:
            pass      

    cloud_db.document().set({
        'phoneNumber': phone_number,
        'location': location
    })

    list_numbers = get_subscribers(user.place)

    return Response({
        'status': 'Successfully Added.',   
        'list_numbers': list_numbers
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def import_subscriber(request):
    username = request.user.username
    
    try:
        user = AuthorizeUser.objects.get(username=username)
    except AuthorizeUser.DoesNotExist:
        return Response({
            'status': 'Authorize User Does Not Exist.'
        }, status=status.HTTP_400_BAD_REQUEST)

    workbook = load_workbook(filename=request.data['file'].file, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    REQUIRED_HEADERS = ['Phone Number', 'Location']
    excel_headers = list()
    
    cloud_db = db.collection('Subscribers')

    for index, row in enumerate(worksheet.values):
        if index == 0:
            for header in row:
                excel_headers.append(header)

            for header in excel_headers:
                if header not in REQUIRED_HEADERS:
                    return Response({
                        'status': 'This header ({}) is required.'.format(header)
                    }, status=status.HTTP_400_BAD_REQUEST)
        else:
            if row[0] != None and row[1] != None:
                phone_number = '+' + str(row[0])
                location = row[1]

                is_exist = False
                for cloud in cloud_db.get():
                    try:
                        if cloud.get('phoneNumber') == phone_number:
                            is_exist = True
                    except KeyError:
                        pass    

                if not is_exist:
                    cloud_db.document().set({
                        'phoneNumber': phone_number,
                        'location': location
                    })

    list_numbers = get_subscribers(user.place)
    
    return Response({
        'status': 'Successfully Imported.',
        'list_numbers': list_numbers
    },status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def edit_subscriber(request):
    username = request.user.username
    
    try:
        user = AuthorizeUser.objects.get(username=username)
    except AuthorizeUser.DoesNotExist:
        return Response({
            'status': 'Authorize User Does Not Exist.'
        }, status=status.HTTP_400_BAD_REQUEST)

    old_phone_number = request.data['old_phone_number']
    new_phone_number = request.data['new_phone_number']
    old_location = request.data['old_location']
    location = request.data['location']

    cloud_db = db.collection('Subscribers')

    for cloud in cloud_db.get():
        try:
            if cloud.get('phoneNumber') == new_phone_number:
                return Response({
                    'status': 'This number already exist!'
                }, status=status.HTTP_400_BAD_REQUEST)
        except KeyError:
            pass    

    for cloud in cloud_db.get():
        try:
            if old_phone_number == cloud.get('phoneNumber'):
                doc_id = cloud.id
                cloud_db.document(doc_id).set({
                    'phoneNumber': new_phone_number,
                    'location': location
                })
        except KeyError:
            pass

    list_numbers = get_subscribers(user.place)

    return Response({
        'status': 'Updated Successfully',
        'list_numbers': list_numbers
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def delete_subscriber(request):
    username = request.user.username
    
    try:
        user = AuthorizeUser.objects.get(username=username)
    except AuthorizeUser.DoesNotExist:
        return Response({
            'status': 'Authorize User Does Not Exist.'
        }, status=status.HTTP_400_BAD_REQUEST)

    phone_number = request.data['phone_number']
    location = request.data['location']

    cloud_db = db.collection('Subscribers')
    for cloud in cloud_db.get():
        try:
            if phone_number == cloud.get('phoneNumber'):
                doc_id = cloud.id
                cloud_db.document(doc_id).delete()

        except KeyError:
            pass
    
    list_numbers = get_subscribers(user.place)

    return Response({
        'status': 'Deleted Successfully',
        'list_numbers': list_numbers
    }, status=status.HTTP_200_OK)

def _get_not_approve(username):
    users = AuthorizeUser.objects.filter(
        is_authorize=False, is_complete_approval=False).exclude(username=username)

    serializer = AuthorizeUserSerializer(users, many=True)

    return serializer.data


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_approval(request):
    username = request.user.username
    data = _get_not_approve(username)

    return Response(data, status=status.HTTP_200_OK)    

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_admin_data(request):
    username = request.user.username
    approval = _get_not_approve(username)
    
    # users = AuthorizeUser.objects.all()
    # user_serializer = AuthorizeUserSerializer(users, many=True)

    list_numbers = admin_subscribers()

    announcements = Announcements.objects.all()
    announcement_serializer = AnnouncementSerializer(announcements, many=True)

    return Response({
        'approval': approval,
        'list_numbers': list_numbers,
        'announcements': announcement_serializer.data
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def authorize_approval(request):
    username = request.user.username
    
    try:
        user_id = request.data['user_id']
    except KeyError:
        return Response({
            'user_id': 'This field is required.'
        }, status=status.HTTP_400_BAD_REQUEST)

    user = AuthorizeUser.objects.get(id=user_id)
    user.is_authorize = True
    user.is_complete_approval = True
    user.save()

    data = _get_not_approve(username)

    return Response(data, status=status.HTTP_200_OK)

@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def authorize_disapproval(request):
    username = request.user.username
    
    try:
        user_id = request.data['user_id']
    except KeyError:
        return Response({
            'user_id': 'This field is required.'
        }, status=status.HTTP_400_BAD_REQUEST)

    user = AuthorizeUser.objects.get(id=user_id)
    user.is_complete_approval = True
    user.save()

    data = _get_not_approve(username)

    return Response(data, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def authorize_announcements(request):
    announcements = Announcements.objects.all()
    serializer = AnnouncementSerializer(announcements, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)
